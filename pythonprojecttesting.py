  
#using pandas library
import pandas as pd
#using openpyxl to write in master sheet
from openpyxl import load_workbook

xls = pd.ExcelFile('PythonExcelSheets.xlsx')
df1 = pd.read_excel(xls, 'Sheet1')
df2 = pd.read_excel(xls, 'Sheet2')
df3 = pd.read_excel(xls, 'Sheet3')
df4 = pd.read_excel(xls, 'Sheet4')
df5 = pd.read_excel(xls, 'Sheet5')
print("Type Print", "\n")

df = pd.merge(df1, df2)
df = pd.merge(df, df3)
df = pd.merge(df, df4)
df = pd.merge(df, df5)
print(df, "\n")

path = r"PythonExcelSheets.xlsx"
print("Enter Unique Id Ps No")
x = int(input())
df1 = pd.DataFrame(df, columns=["Ps No","Name", "Email", "Dept ID", "Module Name","Location","Domain", "Duration of Internship","Floor", "Stipend","Gender","Age","Phone","Education","Occupation","Experience (Years)","Salary","Company Name","Address","City","County","State","ZIP","Degree","Unnamed: 10","Invigilator Name","Subject","Invigilator  Mail ID",
                               "Mobile numberof the Invigilator", "10th Marks", "12th Marks", "Virtual Room No", "Product","Discount Band", "Units Sold","Manufacturing Price","Sale Price","Gross Sales","Profit"])
df1.set_index("Ps No",inplace=True)
result = df1.loc[x]
print(result)


ExcelWorkbook = load_workbook('PythonExcelSheets.xlsx')
writer = pd.ExcelWriter('PythonExcelSheets.xlsx', engine='openpyxl')

writer.book = ExcelWorkbook
if 'mastersheet' in ExcelWorkbook.sheetnames:
    pfd = ExcelWorkbook['mastersheet']
    ExcelWorkbook.remove(pfd)
result.to_excel(writer, sheet_name='mastersheet')

writer.save()
writer.close()
